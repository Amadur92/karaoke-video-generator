#!/usr/bin/env python3
from __future__ import annotations

import argparse
import posixpath
import subprocess
import tempfile
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook


NS = {
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "p14": "http://schemas.microsoft.com/office/powerpoint/2010/main",
}

for prefix in ("p", "a", "r", "p14"):
    etree.register_namespace(prefix, NS[prefix])


VIDEO_OFF = {"x": "623888", "y": "3795870"}
VIDEO_EXT = {"cx": "10944225", "cy": "1818337"}
VIDEO_CROP = "4545"


def q(ns: str, tag: str) -> str:
    return f"{{{NS[ns]}}}{tag}"


def serialize(root: etree._Element) -> bytes:
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def resolve_part(source_part: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def load_tracks(xlsx_path: Path, media_dir: Path) -> list[dict]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    tracks = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            number = int(row[0])
        except Exception:
            continue
        media = media_dir / f"media{number}.mp4"
        if not media.exists():
            raise FileNotFoundError(f"Missing video for #{number}: {media}")
        tracks.append(
            {
                "num": number,
                "artist": str(row[1] or "").strip(),
                "title": str(row[2] or "").strip(),
                "media": media,
            }
        )
    tracks.sort(key=lambda item: item["num"])
    actual = [item["num"] for item in tracks]
    expected = list(range(1, 101))
    if actual != expected:
        raise RuntimeError(f"Expected tracks 1..100, got {actual[:5]}...{actual[-5:]}")
    return tracks


def font_size_for(artist: str, title: str) -> str:
    longest = max(len(artist), len(title))
    total = len(artist) + len(title)
    if longest > 46 or total > 76:
        return "2200"
    if longest > 38 or total > 64:
        return "2500"
    if longest > 30 or total > 54:
        return "2800"
    return "3200"


def clear_extra_texts(nodes: list[etree._Element], start: int = 1) -> None:
    for text_node in nodes[start:]:
        text_node.text = ""


def normalize_title_run_style(shape: etree._Element, size: str) -> None:
    for run_props in shape.xpath(".//a:rPr", namespaces=NS):
        run_props.set("lang", "ru-RU")
        run_props.set("sz", size)
        run_props.set("b", "1")
        run_props.set("dirty", "0")
        for child in list(run_props):
            if child.tag in (q("a", "solidFill"), q("a", "latin"), q("a", "cs")):
                run_props.remove(child)
        solid_fill = etree.Element(q("a", "solidFill"))
        etree.SubElement(solid_fill, q("a", "schemeClr"), val="bg1")
        run_props.insert(0, solid_fill)
        run_props.append(etree.Element(q("a", "latin"), typeface="Montserrat", pitchFamily="2", charset="-52"))
        run_props.append(etree.Element(q("a", "cs"), typeface="TT Commons Bold"))


def set_paragraph_text(paragraph: etree._Element, value: str) -> None:
    paragraph_props = paragraph.xpath("./a:pPr", namespaces=NS)
    if paragraph_props:
        paragraph_props[0].set("algn", "ctr")
    else:
        paragraph.insert(0, etree.Element(q("a", "pPr"), algn="ctr"))
    texts = paragraph.xpath(".//a:t", namespaces=NS)
    if not texts:
        run = etree.SubElement(paragraph, q("a", "r"))
        etree.SubElement(run, q("a", "rPr"), lang="ru-RU", sz="3200", b="1")
        texts = [etree.SubElement(run, q("a", "t"))]
    texts[0].text = value
    clear_extra_texts(texts)


def find_track_title_shape(slide_root: etree._Element) -> etree._Element:
    candidates = []
    for shape in slide_root.xpath("//p:sp", namespaces=NS):
        name = shape.xpath(".//p:cNvPr[1]/@name", namespaces=NS)
        texts = [t.strip() for t in shape.xpath(".//a:t/text()", namespaces=NS) if t.strip()]
        off = shape.xpath(".//a:xfrm/a:off[1]", namespaces=NS)
        ext = shape.xpath(".//a:xfrm/a:ext[1]", namespaces=NS)
        if not texts or not off or not ext:
            continue
        x = int(off[0].get("x", "0"))
        y = int(off[0].get("y", "0"))
        cx = int(ext[0].get("cx", "0"))
        score = 0
        if name and name[0] == "Текстовое поле 9":
            score += 10
        if 2_000_000 <= x <= 3_000_000 and 500_000 <= y <= 900_000:
            score += 5
        if cx > 5_000_000:
            score += 3
        candidates.append((score, shape))
    if not candidates:
        raise RuntimeError("Track slide title shape not found")
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0][1]


def update_track_title(slide_root: etree._Element, track: dict) -> None:
    shape = find_track_title_shape(slide_root)
    artist = track["artist"].upper()
    title = track["title"].upper()
    paragraphs = shape.xpath(".//a:p", namespaces=NS)
    if len(paragraphs) < 2:
        tx_body = shape.xpath("./p:txBody", namespaces=NS)[0]
        paragraphs.append(etree.SubElement(tx_body, q("a", "p")))
    set_paragraph_text(paragraphs[0], f"{artist} –")
    set_paragraph_text(paragraphs[1], title)
    for paragraph in paragraphs[2:]:
        parent = paragraph.getparent()
        if parent is not None:
            parent.remove(paragraph)
    size = font_size_for(artist, title)
    normalize_title_run_style(shape, size)


def update_video_picture(slide_root: etree._Element) -> str:
    pictures = slide_root.xpath(
        '//p:pic[.//a:videoFile or .//p14:media]',
        namespaces=NS,
    )
    if not pictures:
        raise RuntimeError("Video picture not found")
    picture = pictures[0]
    xfrm = picture.xpath(".//a:xfrm", namespaces=NS)[0]
    off = xfrm.xpath("./a:off", namespaces=NS)[0]
    ext = xfrm.xpath("./a:ext", namespaces=NS)[0]
    off.attrib.update(VIDEO_OFF)
    ext.attrib.update(VIDEO_EXT)

    blip_fill = picture.xpath("./p:blipFill", namespaces=NS)[0]
    src_rect = blip_fill.xpath("./a:srcRect", namespaces=NS)
    if src_rect:
        rect = src_rect[0]
    else:
        blip = blip_fill.xpath("./a:blip", namespaces=NS)[0]
        rect = etree.Element(q("a", "srcRect"))
        blip.addnext(rect)
    rect.attrib.update({"l": VIDEO_CROP, "t": VIDEO_CROP, "r": VIDEO_CROP, "b": VIDEO_CROP})

    blip = picture.xpath("./p:blipFill/a:blip", namespaces=NS)[0]
    poster_rid = blip.get(q("r", "embed")) or blip.get(q("r", "link"))
    if not poster_rid:
        raise RuntimeError("Video poster relationship not found")
    return poster_rid


def update_main_table(slide_root: etree._Element, tracks: list[dict]) -> None:
    artists = [track["artist"].upper() for track in tracks]
    tables = slide_root.xpath("//p:graphicFrame", namespaces=NS)
    cells = []
    for table in tables:
        table_name = table.xpath(".//p:cNvPr[1]/@name", namespaces=NS)
        if table_name and table_name[0] not in ("Таблица 26", "Таблица 27"):
            continue
        cells.extend(table.xpath(".//a:tc", namespaces=NS))
    if len(cells) < 100:
        raise RuntimeError(f"Expected at least 100 main-table cells, got {len(cells)}")
    for cell, artist in zip(cells[:100], artists):
        texts = cell.xpath(".//a:t", namespaces=NS)
        if texts:
            texts[0].text = artist
            for extra in texts[1:]:
                extra.text = ""


def read_rels(zip_file: zipfile.ZipFile, slide_index: int) -> dict[str, str]:
    rel_path = f"ppt/slides/_rels/slide{slide_index}.xml.rels"
    rel_root = etree.fromstring(zip_file.read(rel_path))
    return {
        rel.get("Id"): rel.get("Target")
        for rel in rel_root.xpath("//rel:Relationship", namespaces=NS)
    }


def generate_poster(video_path: Path, out_path: Path) -> None:
    base_cmd = [
        "ffmpeg",
        "-y",
        "-hide_banner",
        "-loglevel",
        "error",
        "-ss",
        "3",
        "-i",
        str(video_path),
        "-frames:v",
        "1",
        "-vf",
        "scale=1280:-1",
        str(out_path),
    ]
    result = subprocess.run(base_cmd)
    if result.returncode != 0 or not out_path.exists() or out_path.stat().st_size < 1000:
        subprocess.run(
            [
                "ffmpeg",
                "-y",
                "-hide_banner",
                "-loglevel",
                "error",
                "-i",
                str(video_path),
                "-frames:v",
                "1",
                "-vf",
                "scale=1280:-1",
                str(out_path),
            ],
            check=True,
        )


def build(base_pptx: Path, xlsx_path: Path, media_dir: Path, output_pptx: Path) -> None:
    tracks = load_tracks(xlsx_path, media_dir)
    output_pptx.parent.mkdir(parents=True, exist_ok=True)

    modified_xml: dict[str, bytes] = {}
    replacement_files: dict[str, Path] = {}

    with zipfile.ZipFile(base_pptx, "r") as zin, tempfile.TemporaryDirectory() as temp_name:
        temp_dir = Path(temp_name)

        main_root = etree.fromstring(zin.read("ppt/slides/slide21.xml"))
        update_main_table(main_root, tracks)
        modified_xml["ppt/slides/slide21.xml"] = serialize(main_root)

        for track in tracks:
            slide_index = track["num"] + 21
            slide_path = f"ppt/slides/slide{slide_index}.xml"
            root = etree.fromstring(zin.read(slide_path))
            update_track_title(root, track)
            poster_rid = update_video_picture(root)
            modified_xml[slide_path] = serialize(root)

            rels = read_rels(zin, slide_index)
            media_part = resolve_part(slide_path, f"../media/media{track['num'] + 1}.mp4")
            replacement_files[media_part] = track["media"]

            poster_target = rels.get(poster_rid)
            if not poster_target:
                raise RuntimeError(f"Poster target missing for slide {slide_index}")
            poster_part = resolve_part(slide_path, poster_target)
            poster_path = temp_dir / f"poster{track['num']}.png"
            generate_poster(track["media"], poster_path)
            replacement_files[poster_part] = poster_path

            if track["num"] % 10 == 0:
                print(f"[pptx] prepared {track['num']}/100")

        tmp_output = output_pptx.with_suffix(output_pptx.suffix + ".tmp")
        if tmp_output.exists():
            tmp_output.unlink()
        with zipfile.ZipFile(tmp_output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1, allowZip64=True) as zout:
            for info in zin.infolist():
                if info.filename in modified_xml:
                    zout.writestr(info, modified_xml[info.filename])
                elif info.filename in replacement_files:
                    zout.write(replacement_files[info.filename], info.filename)
                else:
                    with zin.open(info, "r") as source:
                        zout.writestr(info, source.read())
        tmp_output.replace(output_pptx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a Muzloto deck by replacing content in a 175-base PPTX.")
    parser.add_argument("--base", type=Path, required=True)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--media-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    build(args.base, args.xlsx, args.media_dir, args.output)
    print(f"[pptx] wrote {args.output}")


if __name__ == "__main__":
    main()
