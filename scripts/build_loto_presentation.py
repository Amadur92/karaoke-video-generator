#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook


NS = {
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "ct": "http://schemas.openxmlformats.org/package/2006/content-types",
    "ep": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties",
}

for prefix in ("p", "a", "r"):
    etree.register_namespace(prefix, NS[prefix])


def q(ns: str, tag: str) -> str:
    return f"{{{NS[ns]}}}{tag}"


def parse_xml(zf: zipfile.ZipFile, name: str) -> etree._Element:
    return etree.fromstring(zf.read(name))


def serialize(root: etree._Element) -> bytes:
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def load_tracks(xlsx_path: Path, media_dir: Path) -> list[dict]:
    workbook = load_workbook(xlsx_path, read_only=False, data_only=True)
    worksheet = workbook.active
    tracks = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            number = int(row[0])
        except Exception:
            continue
        artist = str(row[1] or "").strip()
        title = str(row[2] or "").strip()
        media = media_dir / f"media{number}.mp4"
        if not media.exists():
            raise FileNotFoundError(f"Missing video for #{number}: {media}")
        tracks.append({"num": number, "artist": artist, "title": title, "media": media})
    tracks.sort(key=lambda item: item["num"])
    expected = list(range(1, len(tracks) + 1))
    actual = [item["num"] for item in tracks]
    if actual != expected:
        raise RuntimeError(f"Expected continuous numbers {expected[:3]}...{expected[-3:]}, got {actual}")
    return tracks


def probe_video_aspect(video_path: Path) -> float:
    output = subprocess.check_output(
        [
            "ffprobe",
            "-v",
            "error",
            "-select_streams",
            "v:0",
            "-show_entries",
            "stream=width,height",
            "-of",
            "csv=p=0:s=x",
            str(video_path),
        ],
        text=True,
    ).strip()
    width_text, height_text = output.split("x", 1)
    return int(width_text) / int(height_text)


def generate_posters(tracks: list[dict], poster_dir: Path) -> None:
    poster_dir.mkdir(parents=True, exist_ok=True)
    for item in tracks:
        poster = poster_dir / f"poster{item['num']}.jpg"
        if poster.exists() and poster.stat().st_size > 1000:
            continue
        cmd = [
            "ffmpeg",
            "-y",
            "-hide_banner",
            "-loglevel",
            "error",
            "-ss",
            "3",
            "-i",
            str(item["media"]),
            "-frames:v",
            "1",
            "-vf",
            "scale=1280:-1",
            "-q:v",
            "5",
            str(poster),
        ]
        result = subprocess.run(cmd)
        if result.returncode != 0 or not poster.exists() or poster.stat().st_size < 1000:
            subprocess.run(
                [
                    "ffmpeg",
                    "-y",
                    "-hide_banner",
                    "-loglevel",
                    "error",
                    "-i",
                    str(item["media"]),
                    "-frames:v",
                    "1",
                    "-vf",
                    "scale=1280:-1",
                    "-q:v",
                    "5",
                    str(poster),
                ],
                check=True,
            )
        if item["num"] % 10 == 0:
            print(f"[pptx] posters {item['num']}/{len(tracks)}")


def font_size_for(artist: str, title: str) -> str:
    max_len = max(len(artist), len(title))
    total = len(artist) + len(title)
    if max_len > 50 or total > 72:
        return "2200"
    if max_len > 40 or total > 62:
        return "2500"
    if max_len > 32 or total > 54:
        return "2800"
    return "3200"


def set_title(root: etree._Element, artist: str, title: str) -> None:
    shape = root.xpath('//p:sp[p:nvSpPr/p:cNvPr/@id="12"]', namespaces=NS)[0]
    paragraphs = shape.xpath(".//a:p", namespaces=NS)
    first_texts = paragraphs[0].xpath(".//a:t", namespaces=NS)
    while len(first_texts) < 3:
        run = etree.SubElement(paragraphs[0], q("a", "r"))
        etree.SubElement(run, q("a", "rPr"), lang="ru-RU", sz="3200", b="1")
        etree.SubElement(run, q("a", "t"))
        first_texts = paragraphs[0].xpath(".//a:t", namespaces=NS)
    first_texts[0].text = artist.upper()
    first_texts[1].text = " "
    first_texts[2].text = "–"
    for extra in first_texts[3:]:
        extra.text = ""

    second_texts = paragraphs[1].xpath(".//a:t", namespaces=NS)
    if not second_texts:
        run = etree.SubElement(paragraphs[1], q("a", "r"))
        etree.SubElement(run, q("a", "rPr"), lang="ru-RU", sz="3200", b="1")
        etree.SubElement(run, q("a", "t"))
        second_texts = paragraphs[1].xpath(".//a:t", namespaces=NS)
    second_texts[0].text = title.upper()
    for extra in second_texts[1:]:
        extra.text = ""

    size = font_size_for(artist.upper(), title.upper())
    if size != "3200":
        for run_props in shape.xpath(".//a:rPr", namespaces=NS):
            run_props.set("sz", size)


def make_number_shape(number: int) -> etree._Element:
    shape = etree.Element(q("p", "sp"))
    nv = etree.SubElement(shape, q("p", "nvSpPr"))
    etree.SubElement(nv, q("p", "cNvPr"), id="1200", name="Slide Number Overlay")
    etree.SubElement(nv, q("p", "cNvSpPr"), txBox="1")
    etree.SubElement(nv, q("p", "nvPr"))

    sppr = etree.SubElement(shape, q("p", "spPr"))
    xfrm = etree.SubElement(sppr, q("a", "xfrm"))
    etree.SubElement(xfrm, q("a", "off"), x="10236000", y="410000")
    etree.SubElement(xfrm, q("a", "ext"), cx="1470000", cy="1470000")
    geom = etree.SubElement(sppr, q("a", "prstGeom"), prst="ellipse")
    etree.SubElement(geom, q("a", "avLst"))
    fill = etree.SubElement(sppr, q("a", "solidFill"))
    etree.SubElement(fill, q("a", "srgbClr"), val="2C2D83")
    line = etree.SubElement(sppr, q("a", "ln"), w="10500")
    line_fill = etree.SubElement(line, q("a", "solidFill"))
    line_color = etree.SubElement(line_fill, q("a", "srgbClr"), val="FFFFFF")
    etree.SubElement(line_color, q("a", "alpha"), val="85000")

    tx = etree.SubElement(shape, q("p", "txBody"))
    etree.SubElement(tx, q("a", "bodyPr"), rtlCol="0", anchor="ctr")
    etree.SubElement(tx, q("a", "lstStyle"))
    para = etree.SubElement(tx, q("a", "p"))
    etree.SubElement(para, q("a", "pPr"), algn="ctr")
    run = etree.SubElement(para, q("a", "r"))
    run_props = etree.SubElement(run, q("a", "rPr"), lang="en-US", sz="5200", b="1")
    text_fill = etree.SubElement(run_props, q("a", "solidFill"))
    etree.SubElement(text_fill, q("a", "srgbClr"), val="FFFFFF")
    etree.SubElement(run_props, q("a", "latin"), typeface="Montserrat")
    etree.SubElement(run_props, q("a", "cs"), typeface="Arial")
    etree.SubElement(run, q("a", "t")).text = str(number)
    return shape


def add_number_overlay(root: etree._Element, number: int) -> None:
    sp_tree = root.xpath("//p:cSld/p:spTree", namespaces=NS)[0]
    children = list(sp_tree)
    insert_at = len(children)
    for index, child in enumerate(children):
        cnv = child.xpath(".//p:cNvPr[1]", namespaces=NS)
        if cnv and cnv[0].get("id") == "9":
            insert_at = index + 1
            break
    sp_tree.insert(insert_at, make_number_shape(number))


def fit_video_shape(root: etree._Element, aspect: float, zoom: float) -> None:
    # Keep the shape inside the white rounded card from the template, but crop
    # inside it to make the lyrics larger without stretching.
    frame_x = 623_888
    frame_cx = 10_944_225
    card_y = 3_115_773
    card_cy = 3_178_532
    frame_cy = round(frame_cx / aspect)
    frame_y = round(card_y + (card_cy - frame_cy) / 2)
    crop_each = str(round(((zoom - 1.0) / (2.0 * zoom)) * 100_000))

    pic = root.xpath('//p:pic[p:nvPicPr/p:cNvPr[starts-with(@name,"media")]]', namespaces=NS)[0]
    xfrm = pic.xpath("./p:spPr/a:xfrm", namespaces=NS)[0]
    off = xfrm.find(q("a", "off"))
    ext = xfrm.find(q("a", "ext"))
    off.set("x", str(frame_x))
    off.set("y", str(frame_y))
    ext.set("cx", str(frame_cx))
    ext.set("cy", str(frame_cy))

    src_rect = pic.xpath("./p:blipFill/a:srcRect", namespaces=NS)
    if src_rect:
        src = src_rect[0]
    else:
        blip_fill = pic.xpath("./p:blipFill", namespaces=NS)[0]
        stretch = blip_fill.find(q("a", "stretch"))
        src = etree.Element(q("a", "srcRect"))
        insert_at = list(blip_fill).index(stretch) if stretch is not None else len(blip_fill)
        blip_fill.insert(insert_at, src)
    src.set("l", crop_each)
    src.set("t", crop_each)
    src.set("r", crop_each)
    src.set("b", crop_each)


def build_deck(
    template: Path,
    xlsx: Path,
    media_dir: Path,
    output: Path,
    zoom: float,
    poster_dir: Path | None,
) -> None:
    tracks = load_tracks(xlsx, media_dir)
    aspect = probe_video_aspect(tracks[0]["media"])
    print(f"[pptx] tracks={len(tracks)} video_aspect={aspect:.3f} zoom={zoom:.3f}")

    scratch = poster_dir or Path(tempfile.mkdtemp(prefix="loto_pptx_posters_"))
    generate_posters(tracks, scratch)

    with zipfile.ZipFile(template, "r") as zin:
        slide_template = parse_xml(zin, "ppt/slides/slide1.xml")
        rels_template = parse_xml(zin, "ppt/slides/_rels/slide1.xml.rels")
        pres = parse_xml(zin, "ppt/presentation.xml")
        pres_rels = parse_xml(zin, "ppt/_rels/presentation.xml.rels")
        ctypes = parse_xml(zin, "[Content_Types].xml")
        app = parse_xml(zin, "docProps/app.xml")

        slide_id_list = pres.find(q("p", "sldIdLst"))
        for child in list(slide_id_list):
            slide_id_list.remove(child)
        for rel in list(pres_rels):
            if rel.get("Type") == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide":
                pres_rels.remove(rel)
        for index in range(1, len(tracks) + 1):
            rel_id = f"rId{100 + index}"
            etree.SubElement(slide_id_list, q("p", "sldId"), id=str(688 + index), **{q("r", "id"): rel_id})
            etree.SubElement(
                pres_rels,
                "Relationship",
                Id=rel_id,
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide",
                Target=f"slides/slide{index}.xml",
            )

        existing_defaults = {el.get("Extension") for el in ctypes.findall(q("ct", "Default"))}
        if "jpg" not in existing_defaults:
            etree.SubElement(ctypes, q("ct", "Default"), Extension="jpg", ContentType="image/jpeg")
        existing_overrides = {el.get("PartName") for el in ctypes.findall(q("ct", "Override"))}
        slide_content_type = "application/vnd.openxmlformats-officedocument.presentationml.slide+xml"
        for index in range(1, len(tracks) + 1):
            part_name = f"/ppt/slides/slide{index}.xml"
            if part_name not in existing_overrides:
                etree.SubElement(ctypes, q("ct", "Override"), PartName=part_name, ContentType=slide_content_type)

        for tag, value in (("Slides", str(len(tracks))), ("MMClips", str(len(tracks)))):
            element = app.find(q("ep", tag))
            if element is not None:
                element.text = value

        tmp_output = output.with_suffix(".tmp.pptx")
        with zipfile.ZipFile(tmp_output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1) as zout:
            skip = {"ppt/presentation.xml", "ppt/_rels/presentation.xml.rels", "[Content_Types].xml", "docProps/app.xml"}
            for info in zin.infolist():
                name = info.filename
                if name in skip:
                    continue
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name):
                    continue
                if re.fullmatch(r"ppt/slides/_rels/slide\d+\.xml\.rels", name):
                    continue
                if re.fullmatch(r"ppt/media/media\d+\.mp4", name):
                    continue
                zout.writestr(info, zin.read(name))

            zout.writestr("[Content_Types].xml", serialize(ctypes))
            zout.writestr("ppt/presentation.xml", serialize(pres))
            zout.writestr("ppt/_rels/presentation.xml.rels", serialize(pres_rels))
            zout.writestr("docProps/app.xml", serialize(app))

            for item in tracks:
                number = item["num"]
                root = etree.fromstring(etree.tostring(slide_template))
                set_title(root, item["artist"], item["title"])
                add_number_overlay(root, number)
                fit_video_shape(root, aspect=aspect, zoom=zoom)
                zout.writestr(f"ppt/slides/slide{number}.xml", serialize(root))

                rels = etree.fromstring(etree.tostring(rels_template))
                for rel in rels.findall(q("rel", "Relationship")):
                    if rel.get("Id") in ("rId1", "rId2"):
                        rel.set("Target", f"../media/media{number}.mp4")
                    elif rel.get("Id") == "rId6":
                        rel.set("Target", f"../media/poster{number}.jpg")
                zout.writestr(f"ppt/slides/_rels/slide{number}.xml.rels", serialize(rels))
                zout.write(item["media"], f"ppt/media/media{number}.mp4")
                zout.write(scratch / f"poster{number}.jpg", f"ppt/media/poster{number}.jpg")

        os.replace(tmp_output, output)
        print(f"[pptx] wrote {output}")
        print(f"[pptx] size={output.stat().st_size / 1024 / 1024:.1f} MB")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a Muzloto PowerPoint deck from a one-slide template.")
    parser.add_argument("--template", required=True, type=Path, help="One-slide PPTX template.")
    parser.add_argument("--xlsx", required=True, type=Path, help="Track list with columns: number, artist, title.")
    parser.add_argument("--media-dir", required=True, type=Path, help="Folder with mediaN.mp4 files.")
    parser.add_argument("--output", required=True, type=Path, help="Output PPTX path.")
    parser.add_argument("--zoom", type=float, default=1.10, help="Internal video crop zoom. 1.10 crops 4.545%% per edge.")
    parser.add_argument("--poster-dir", type=Path, default=None, help="Optional folder for generated poster JPGs.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    build_deck(
        template=args.template,
        xlsx=args.xlsx,
        media_dir=args.media_dir,
        output=args.output,
        zoom=args.zoom,
        poster_dir=args.poster_dir,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
